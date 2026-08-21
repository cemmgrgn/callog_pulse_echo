"""History: sessions and certificates.

Deleting a certificate is a "soft" delete: the record doesn't leave the
database, it's only marked. This avoids gaps in the numbering series and
preserves measurement data and the audit trail; only admins can see
deleted records.
"""

import os
import zipfile

from .. import audit, certificate, db, perms, points, sessions, theme, waveform
from ..qt import Qt, QtGui, QtWidgets
from .util import (DateRangeFilter, empty_state, fit_table, PAGE_MARGIN,
                   PAGE_SPACING)

STATUS_TR = {
    "running": "sürüyor", "completed": "tamamlandı", "aborted": "iptal edildi",
    "draft": "taslak",
}


class HistoryPage(QtWidgets.QWidget):

    def __init__(self, app_state, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.state = app_state

        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGIN)
        root.setSpacing(PAGE_SPACING)

        title = QtWidgets.QLabel("Geçmiş kayıtlar")
        title.setProperty("h1", True)
        root.addWidget(title)

        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setObjectName("innerTabs")
        self.tabs.addTab(self._sessions_tab(), "Ölçüm oturumları")
        self.tabs.addTab(self._certificates_tab(), "Sertifikalar")
        self.tabs.currentChanged.connect(lambda _i: self.reload())
        root.addWidget(self.tabs, 1)

    # =====================================================================
    # Sessions
    # =====================================================================
    def _sessions_tab(self):
        w = QtWidgets.QWidget()
        lay = QtWidgets.QVBoxLayout(w)
        lay.setContentsMargins(0, 10, 0, 0)
        lay.addLayout(self._session_filters())

        split = QtWidgets.QSplitter(Qt.Horizontal)
        split.addWidget(self._session_list())
        split.addWidget(self._detail())
        split.setStretchFactor(0, 3)
        split.setStretchFactor(1, 2)
        lay.addWidget(split, 1)
        lay.addLayout(self._session_buttons())
        return w

    def _session_filters(self):
        grid = QtWidgets.QGridLayout()
        self.search = QtWidgets.QLineEdit()
        self.search.setPlaceholderText("Ara: seri no, model, şirket veya operatör")
        self.search.textChanged.connect(self.reload_sessions)

        self.dut_filter = QtWidgets.QComboBox()
        self.dut_filter.currentIndexChanged.connect(self.reload_sessions)
        self.instrument_filter = QtWidgets.QComboBox()
        self.instrument_filter.currentIndexChanged.connect(self.reload_sessions)
        self.status_filter = QtWidgets.QComboBox()
        states = [("Tüm durumlar", None), ("Tamamlandı", "completed"),
                  ("İptal edildi", "aborted"), ("Sürüyor", "running")]
        # Offering this option to a user who can't see deleted records would
        # make them see an empty list on selecting it and wonder "is this
        # broken?".
        if self.state.can(perms.SESSION_VIEW_DELETED):
            states.append(("Silinmiş", "deleted"))
        for label, value in states:
            self.status_filter.addItem(label, value)
        self.status_filter.currentIndexChanged.connect(self.reload_sessions)

        self.only_mine = QtWidgets.QCheckBox("Yalnızca benim ölçümlerim")
        self.only_mine.toggled.connect(self.reload_sessions)

        self.date_filter = DateRangeFilter()
        self.date_filter.changed.connect(self.reload_sessions)

        grid.addWidget(self.search, 0, 0, 1, 3)
        grid.addWidget(self.only_mine, 0, 3)
        grid.addWidget(QtWidgets.QLabel("Kalibre edilen cihaz"), 1, 0)
        grid.addWidget(self.dut_filter, 1, 1)
        grid.addWidget(QtWidgets.QLabel("Referans cihaz"), 1, 2)
        grid.addWidget(self.instrument_filter, 1, 3)
        grid.addWidget(QtWidgets.QLabel("Durum"), 1, 4)
        grid.addWidget(self.status_filter, 1, 5)
        grid.addWidget(QtWidgets.QLabel("Tarih"), 2, 0)
        grid.addWidget(self.date_filter, 2, 1, 1, 5)
        for c in (1, 3, 5):
            grid.setColumnStretch(c, 1)
        return grid

    def _session_list(self):
        self.table = QtWidgets.QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(
            ["#", "Oturum adı", "Tarih", "Cihaz", "Fonksiyon", "Operatör", "Durum"])
        self.table.doubleClicked.connect(self._rename_session)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        # Multi-selection for batch operations and comparison. The detail
        # panel still shows the first selected row.
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.itemSelectionChanged.connect(self._show_detail)
        return self.table

    def _detail(self):
        self.detail = QtWidgets.QTextBrowser()
        self.detail.setOpenExternalLinks(False)
        return self.detail

    def _session_buttons(self):
        """Button row.

        Export and batch operations are grouped into a single menu button:
        a row of six side-by-side buttons was determining the narrowest the
        page could be drawn, and the window didn't fit on 1366px screens.
        """
        row = QtWidgets.QHBoxLayout()
        self.cert_btn = QtWidgets.QPushButton("Sertifika üret (PDF)")
        self.cert_btn.setProperty("primary", True)
        self.cert_btn.clicked.connect(self._make_certificate)

        self.export_btn = QtWidgets.QToolButton()
        self.export_btn.setText("Dışa aktar")
        self.export_btn.setPopupMode(QtWidgets.QToolButton.InstantPopup)
        menu = QtWidgets.QMenu(self.export_btn)
        self.act_docx = menu.addAction("Word'e aktar (.docx)")
        self.act_docx.triggered.connect(self._export_docx)
        self.act_xlsx = menu.addAction("Excel'e aktar (.xlsx)")
        self.act_xlsx.triggered.connect(self._export_excel)
        menu.addSeparator()
        self.act_batch_xlsx = menu.addAction("Seçilenleri Excel'e aktar…")
        self.act_batch_xlsx.triggered.connect(self._batch_excel)
        self.act_batch_cert = menu.addAction("Seçilenlere sertifika üret…")
        self.act_batch_cert.triggered.connect(self._batch_certificates)
        self.export_btn.setMenu(menu)
        self.export_btn.setToolTip(
            "Tek oturum için dışa aktarma ve seçili oturumlar için toplu "
            "işlem.\nÇoklu seçim: Ctrl ile tek tek, Shift ile aralık.")

        self.compare_btn = QtWidgets.QPushButton("Karşılaştır")
        self.compare_btn.setToolTip(
            "Seçili oturumların okumalarını aynı grafikte üst üste çizer.\n"
            "En az iki oturum seçin.")
        self.compare_btn.clicked.connect(self._compare)

        self.rename_btn = QtWidgets.QPushButton("Yeniden adlandır")
        self.rename_btn.clicked.connect(self._rename_session)
        self.delete_session_btn = QtWidgets.QPushButton("Oturumu sil")
        self.delete_session_btn.setProperty("danger", True)
        self.delete_session_btn.clicked.connect(self._delete_session)
        self.restore_session_btn = QtWidgets.QPushButton("Geri al")
        self.restore_session_btn.clicked.connect(self._restore_session)
        for b in (self.cert_btn, self.export_btn, self.compare_btn,
                  self.rename_btn, self.delete_session_btn,
                  self.restore_session_btn):
            row.addWidget(b)
        # For a user without permission the button isn't grayed out, it's
        # gone: a permanently disabled button makes the user wonder "why
        # doesn't this work for me?".
        self.delete_session_btn.setVisible(self.state.can(perms.SESSION_DELETE))
        self.restore_session_btn.setVisible(self.state.can(perms.SESSION_RESTORE))
        row.addStretch(1)
        return row

    # --- data ---------------------------------------------------------------
    def _reload_filter_combos(self):
        """Refreshes the filter dropdowns, preserving the current selection."""
        for combo, sql, label_fn, all_text in (
            (self.dut_filter,
             "SELECT d.id, d.manufacturer, d.model, d.serial_no FROM duts d"
             " JOIN sessions s ON s.dut_id = d.id GROUP BY d.id"
             " ORDER BY d.manufacturer, d.model",
             lambda r: "%s %s — %s" % (r["manufacturer"], r["model"], r["serial_no"]),
             "Tüm cihazlar"),
            (self.instrument_filter,
             "SELECT i.id, i.brand, i.model, i.serial_no FROM instruments i"
             " ORDER BY i.id",
             lambda r: "%s %s — %s" % (r["brand"], r["model"], r["serial_no"]),
             "Tüm referanslar"),
        ):
            previous = combo.currentData()
            combo.blockSignals(True)
            combo.clear()
            combo.addItem(all_text, None)
            for r in db.query(sql):
                combo.addItem(label_fn(r), r["id"])
            index = combo.findData(previous)
            combo.setCurrentIndex(index if index >= 0 else 0)
            combo.blockSignals(False)

    def reload(self):
        self._reload_filter_combos()
        self.reload_sessions()
        self.reload_certificates()

    # --- focus requests from outside (global search) -----------------------
    def _clear_session_filters(self):
        """Resets the filters — so the searched-for record isn't hidden by them."""
        self.search.blockSignals(True)
        self.search.clear()
        self.search.blockSignals(False)
        self.only_mine.blockSignals(True)
        self.only_mine.setChecked(False)
        self.only_mine.blockSignals(False)
        for combo in (self.dut_filter, self.instrument_filter,
                      self.status_filter, self.date_filter.combo):
            combo.blockSignals(True)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)

    def focus_session(self, session_id):
        """Finds and selects the session in the list. Returns True if found."""
        self.tabs.setCurrentIndex(0)
        self._clear_session_filters()
        self.reload_sessions()
        for r in range(self.table.rowCount()):
            if self.table.item(r, 0).data(Qt.UserRole) == session_id:
                self.table.clearSelection()
                self.table.selectRow(r)
                self.table.scrollToItem(self.table.item(r, 0))
                return True
        return False

    def focus_certificate(self, cert_no):
        """Finds and selects the certificate in the list. Returns True if found."""
        self.tabs.setCurrentIndex(1)
        self.cert_search.blockSignals(True)
        self.cert_search.setText(cert_no or "")
        self.cert_search.blockSignals(False)
        for combo in (self.cert_dut_filter, self.cert_state_filter,
                      self.cert_result_filter, self.cert_date_filter.combo):
            combo.blockSignals(True)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)
        self.reload_certificates()
        for r in range(self.cert_table.rowCount()):
            if self.cert_table.item(r, 0).text() == cert_no:
                self.cert_table.clearSelection()
                self.cert_table.selectRow(r)
                self.cert_table.scrollToItem(self.cert_table.item(r, 0))
                return True
        return False

    def reload_sessions(self):
        sql = (
            "SELECT s.*, d.company, d.manufacturer, d.model,"
            " d.serial_no AS dut_serial, u.full_name AS operator_name"
            " FROM sessions s"
            " JOIN duts d ON d.id = s.dut_id"
            " JOIN users u ON u.id = s.operator_id"
        )
        where, params = [], []
        status = self.status_filter.currentData()

        # Deleted sessions are visible only to admins, and only while that
        # filter is selected
        if status == "deleted":
            if not self._is_admin():
                where.append("0")
            else:
                where.append("s.deleted_at IS NOT NULL")
        else:
            where.append("s.deleted_at IS NULL")

        term = self.search.text().strip()
        if term:
            like = "%" + term + "%"
            where.append("(d.serial_no LIKE ? OR d.model LIKE ? OR d.company LIKE ?"
                         " OR u.full_name LIKE ? OR s.name LIKE ?)")
            params += [like, like, like, like, like]
        if self.dut_filter.currentData() is not None:
            where.append("s.dut_id = ?")
            params.append(self.dut_filter.currentData())
        if self.instrument_filter.currentData() is not None:
            where.append("s.instrument_id = ?")
            params.append(self.instrument_filter.currentData())
        if status is not None and status != "deleted":
            where.append("s.status = ?")
            params.append(status)
        if self.only_mine.isChecked():
            where.append("s.operator_id = ?")
            params.append(self.state.user["id"])
        start, end = self.date_filter.range()
        if start:
            where.append("s.started_at >= ? AND s.started_at < ?")
            params += [start, end]
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY s.id DESC LIMIT 500"

        c = theme.colors()
        rows = db.query(sql, tuple(params))
        self.table.setRowCount(0)
        for r in rows:
            i = self.table.rowCount()
            self.table.insertRow(i)
            status_text = STATUS_TR.get(r["status"], r["status"])
            if r["deleted_at"]:
                status_text = "SİLİNDİ"
            cells = [
                str(r["id"]), sessions.display_name(r),
                (r["started_at"] or "")[:16].replace("T", " "),
                "%s %s" % (r["model"], r["dut_serial"]),
                r["function"], r["operator_name"], status_text,
            ]
            for col, text in enumerate(cells):
                item = QtWidgets.QTableWidgetItem(text)
                if col == 0:
                    item.setData(Qt.UserRole, r["id"])
                if r["deleted_at"]:
                    item.setForeground(QtGui.QColor(c["text_muted"]))
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)
                elif r["is_simulated"]:
                    item.setForeground(QtGui.QColor(c["warn"]))
                self.table.setItem(i, col, item)
        fit_table(self.table, stretch_column=1)
        empty_state(self.table, "Bu süzgece uyan ölçüm oturumu yok.\nTarih aralığı: %s"
                    % self.date_filter.describe())
        self._show_detail()

    def _selected_session_id(self):
        rows = self.table.selectionModel().selectedRows()
        if not rows:
            return None
        return self.table.item(rows[0].row(), 0).data(Qt.UserRole)

    def _selected_session_ids(self):
        """All selected sessions, in list order."""
        rows = sorted(r.row() for r in self.table.selectionModel().selectedRows())
        return [self.table.item(r, 0).data(Qt.UserRole) for r in rows]

    def _show_detail(self):
        sid = self._selected_session_id()
        row = (db.query_one("SELECT deleted_at FROM sessions WHERE id = ?", (sid,))
               if sid is not None else None)
        deleted = bool(row and row["deleted_at"])
        can_manage = self.state.can(perms.SESSION_DELETE)

        chosen = self._selected_session_ids()
        self.cert_btn.setEnabled(sid is not None and not deleted)
        self.export_btn.setEnabled(bool(chosen))
        for act in (self.act_docx, self.act_xlsx):
            act.setEnabled(sid is not None and not deleted)
        for act in (self.act_batch_xlsx, self.act_batch_cert):
            act.setEnabled(len(chosen) >= 1)
        self.compare_btn.setEnabled(len(chosen) >= 2)
        self.rename_btn.setEnabled(sid is not None and not deleted)
        self.delete_session_btn.setEnabled(
            sid is not None and not deleted and can_manage)
        self.restore_session_btn.setEnabled(deleted and self._is_admin())

        if sid is None:
            c = theme.colors()
            self.detail.setHtml(
                "<p style='color:%s'>Soldaki listeden bir oturum seçin — "
                "ölçüm özeti, çevre koşulları ve sertifika durumu burada "
                "görünecek.</p>" % c["text_muted"])
            return
        try:
            d = certificate.collect(sid)
        except Exception as exc:
            self.detail.setHtml("<p>Detay okunamadı: %s</p>" % exc)
            return

        s, dut, inst = d["session"], d["dut"], d["instrument"]
        unit = s["unit"]
        c = theme.colors()

        def f(v):
            return "—" if v is None else "%.7g %s" % (v, unit)

        cert = db.query_one(
            "SELECT * FROM certificates WHERE session_id = ?", (sid,))

        html = ["<h3>%s</h3>" % sessions.display_name(s),
                "<p style='color:%s'>Oturum #%d</p>" % (c["text_muted"], sid)]
        if s["deleted_at"]:
            deleter = db.query_one("SELECT full_name FROM users WHERE id = ?",
                                   (s["deleted_by"],))
            html.append(
                "<p style='color:%s'><b>Bu oturum silinmiş</b> — %s, %s<br>"
                "Gerekçe: %s</p>" % (
                    c["bad"], (deleter["full_name"] if deleter else "—"),
                    (s["deleted_at"] or "")[:16].replace("T", " "),
                    s["delete_reason"] or "—"))
        if s["is_simulated"]:
            html.append("<p style='color:%s'><b>Simülasyon verisi</b> — üretilecek "
                        "sertifika filigranlı ve SIM- serisinden numaralı olur.</p>"
                        % c["warn"])
        html.append("<b>Kalibre edilen cihaz</b><br>%s / %s %s<br>Seri no: %s<br><br>"
                    % (dut["company"], dut["manufacturer"], dut["model"],
                       dut["serial_no"]))
        html.append("<b>Referans standart</b><br>%s %s (SN %s)<br><br>"
                    % (inst["brand"], inst["model"], inst["serial_no"]))

        # In a multi-point plan each point gets its own block: a single
        # block would only show the first point, and you'd only learn
        # "10 V passes but 100 V doesn't" by opening the certificate.
        for p in d["points"]:
            p_unit = p["unit"]

            def pf(v, u=p_unit):
                return "—" if v is None else "%.7g %s" % (v, u)

            p_tol = ("± %g %s" % (p["tolerance"], p_unit)) if p["tolerance"] else "—"
            title = ("<b>Ölçüm</b>" if not d["multi"]
                     else "<b>%d. nokta — %s</b>" % (p["seq"], points.label(p)))
            html.append("%s<br>Fonksiyon: %s<br>n = %d (dışlanan: %d)<br>"
                        "Nominal: %s<br>Tolerans: %s<br>Kriter: %s<br>"
                        "Ortalama: %s<br>Std sapma: %s<br>u(A tipi): %s<br>"
                        "U (k=2): %s<br>En küçük / en büyük: %s / %s<br>"
                        "Sapma: %s<br>%s<br>"
                        % (title, p["function"], p["n"], p["excluded"],
                           pf(p["nominal"]), p_tol,
                           certificate.CRITERION_TR[p["mode"]], pf(p["mean"]),
                           pf(p["std"]), pf(p["u_a"]), pf(p["U"]), pf(p["min"]),
                           pf(p["max"]), pf(p["deviation"]),
                           ("Sonuç: %s<br>" % certificate.VERDICT_TR[p["result"]]
                            if d["multi"] else "")))
        html.append("<b>Ortam</b><br>%s °C / %s %%RH / %s kPa<br><br>"
                    % (s["env_temp"], s["env_rh"], s["env_pressure"]))
        html.append("<b>Sonuç: %s</b><br><br>" % certificate.VERDICT_TR[d["result"]])
        if s["notes"]:
            html.append("<b>Notlar</b><br>%s<br><br>" % s["notes"])
        if cert:
            state = "silinmiş" if cert["deleted_at"] else (
                "onaylandı" if cert["approved_at"] else "onay bekliyor")
            html.append("<b>Sertifika</b><br>%s — %s<br>"
                        % (cert["cert_no"], state))
        self.detail.setHtml("".join(html))

        self.cert_btn.setEnabled(
            s["status"] == "completed" and cert is None and d["total_n"] > 0
            and not s["deleted_at"])

    # --- session name and deletion -------------------------------------------
    def _rename_session(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        row = db.query_one("SELECT * FROM sessions WHERE id = ?", (sid,))
        if row["deleted_at"]:
            QtWidgets.QMessageBox.information(
                self, "Silinmiş oturum",
                "Silinmiş bir oturum yeniden adlandırılamaz. Önce geri alın.")
            return

        current = row["name"] or sessions.default_name(row["dut_id"],
                                                       row["started_at"])
        name, ok = QtWidgets.QInputDialog.getText(
            self, "Oturumu yeniden adlandır",
            "Oturum adı (boş bırakılırsa varsayılana döner):",
            QtWidgets.QLineEdit.Normal, current)
        if not ok:
            return
        try:
            new_name = sessions.rename(sid, name, self.state.user["id"])
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Adlandırılamadı", str(exc))
            return
        self.reload_sessions()
        self.state.status("Oturum adı: %s" % new_name)

    def _delete_session(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        row = db.query_one("SELECT * FROM sessions WHERE id = ?", (sid,))
        reason, ok = QtWidgets.QInputDialog.getText(
            self, "Oturumu sil",
            "'%s' silinecek.\n\nKayıt veritabanından çıkmaz, ham okumalar ve "
            "denetim izi korunur; yöneticiler görmeye devam eder.\n\n"
            "Gerekçe (zorunlu):" % sessions.display_name(row))
        if not ok:
            return
        try:
            sessions.soft_delete(sid, self.state.user["id"], reason)
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Silinemedi", str(exc))
            return
        self.reload_sessions()
        self.state.status("Oturum #%d silindi olarak işaretlendi." % sid)

    def _restore_session(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        try:
            sessions.restore(sid, self.state.user["id"])
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Geri alınamadı", str(exc))
            return
        self.reload_sessions()
        self.state.status("Oturum #%d geri alındı." % sid)

    # --- actions --------------------------------------------------------------
    def _make_certificate(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        sess = db.query_one("SELECT is_simulated FROM sessions WHERE id = ?", (sid,))
        if sess["is_simulated"]:
            ans = QtWidgets.QMessageBox.question(
                self, "Simülasyon sertifikası",
                "Bu oturum simüle edilmiş bir cihazla alındı.\n\n"
                "Üretilecek belge çapraz 'SİMÜLASYON' filigranı taşıyacak, "
                "SIM- serisinden numara alacak ve resmî sertifika yerine "
                "kullanılamayacak.\n\nDevam edilsin mi?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            if ans != QtWidgets.QMessageBox.Yes:
                return
        try:
            path, cert_no, _result = certificate.build_pdf(
                sid, self.state.user["id"])
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Sertifika üretilemedi", str(exc))
            return
        self.reload()
        QtWidgets.QMessageBox.information(
            self, "Sertifika üretildi",
            "%s\n\nDosya:\n%s\n\nOnay için lab sorumlusuna gönderildi."
            % (cert_no, path))
        _open_file(path)

    def _export_docx(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        cert = db.query_one(
            "SELECT cert_no FROM certificates WHERE session_id = ?", (sid,))
        suggested = "%s.docx" % (cert["cert_no"] if cert else "oturum-%d" % sid)
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Word belgesi olarak kaydet", suggested, "Word belgesi (*.docx)")
        if not path:
            return
        try:
            certificate.write_docx(sid, path)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Aktarılamadı", str(exc))
            return
        self.state.status("Word belgesi kaydedildi: %s" % path)
        _open_file(path)

    def _export_excel(self):
        sid = self._selected_session_id()
        if sid is None:
            return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Excel'e aktar", "oturum-%d.xlsx" % sid, "Excel (*.xlsx)")
        if not path:
            return
        try:
            _write_excel(sid, path)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Aktarılamadı", str(exc))
            return
        audit.log("session.export_excel", user_id=self.state.user["id"],
                  entity="session", entity_id=sid, detail={"path": path})
        self.state.status("Excel'e aktarıldı: %s" % path)

    # --- batch operations -----------------------------------------------------
    def _batch_excel(self):
        """Writes each selected session out as a separate .xlsx.

        We don't merge into a single file: each session has its own summary
        + raw data sheets, and putting ten sessions in one file would
        produce a thirty-sheet workbook nobody could navigate. What's
        usually wanted before an audit is "one file per measurement" anyway.
        """
        ids = self._selected_session_ids()
        if not ids:
            return
        directory = QtWidgets.QFileDialog.getExistingDirectory(
            self, "%d oturum için klasör seçin" % len(ids))
        if not directory:
            return

        written, failed = [], []
        for sid in ids:
            row = db.query_one("SELECT * FROM sessions WHERE id = ?", (sid,))
            if row is None or row["deleted_at"]:
                failed.append("#%s: silinmiş oturum" % sid)
                continue
            name = _safe_filename("oturum-%d-%s" % (sid,
                                                    sessions.display_name(row)))
            path = os.path.join(directory, name + ".xlsx")
            try:
                _write_excel(sid, path)
                written.append(sid)
            except Exception as exc:
                failed.append("#%s: %s" % (sid, exc))

        if written:
            audit.log("session.export_excel_batch",
                      user_id=self.state.user["id"],
                      detail={"sessions": written, "directory": directory})
        message = "%d oturum yazıldı: %s" % (len(written), directory)
        if failed:
            message += "\n\nAktarılamayanlar:\n" + "\n".join(failed)
            QtWidgets.QMessageBox.warning(self, "Bazıları aktarılamadı", message)
        else:
            QtWidgets.QMessageBox.information(self, "Toplu aktarım", message)
        self.state.status("%d oturum Excel'e aktarıldı." % len(written))

    def _batch_certificates(self):
        """Generates certificates for eligible selected sessions.

        Ineligible ones (still running, deleted, no readings, already
        certified) are **skipped, with the reason recorded**: a record
        silently skipped in a batch operation means a certificate someone
        assumes was generated but wasn't.
        """
        ids = self._selected_session_ids()
        if not ids:
            return

        eligible, skipped = [], []
        for sid in ids:
            row = db.query_one("SELECT * FROM sessions WHERE id = ?", (sid,))
            if row is None:
                continue
            if row["deleted_at"]:
                skipped.append("#%d: silinmiş" % sid)
            elif row["status"] != "completed":
                skipped.append("#%d: oturum tamamlanmamış" % sid)
            elif certificate.for_session(sid) is not None:
                skipped.append("#%d: sertifikası zaten var" % sid)
            else:
                try:
                    if certificate.collect(sid)["total_n"] == 0:
                        skipped.append("#%d: kayıtlı okuma yok" % sid)
                        continue
                except Exception as exc:
                    skipped.append("#%d: %s" % (sid, exc))
                    continue
                eligible.append((sid, bool(row["is_simulated"])))

        if not eligible:
            QtWidgets.QMessageBox.information(
                self, "Üretilecek sertifika yok",
                "Seçili oturumların hiçbiri uygun değil:\n\n"
                + "\n".join(skipped))
            return

        simulated = sum(1 for _sid, sim in eligible if sim)
        text = "%d oturum için sertifika üretilecek." % len(eligible)
        if simulated:
            text += ("\n\n%d tanesi simülasyon verisi — o belgeler filigranlı "
                     "ve SIM- serisinden numaralı olacak." % simulated)
        if skipped:
            text += "\n\nAtlanacaklar:\n" + "\n".join(skipped)
        text += "\n\nDevam edilsin mi?"
        ans = QtWidgets.QMessageBox.question(
            self, "Toplu sertifika", text,
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if ans != QtWidgets.QMessageBox.Yes:
            return

        made, failed = [], []
        for sid, _sim in eligible:
            try:
                _path, cert_no, _result = certificate.build_pdf(
                    sid, self.state.user["id"])
                made.append(cert_no)
            except Exception as exc:
                failed.append("#%d: %s" % (sid, exc))

        self.reload()
        message = "%d sertifika üretildi:\n%s" % (len(made), ", ".join(made))
        if failed:
            message += "\n\nÜretilemeyenler:\n" + "\n".join(failed)
            QtWidgets.QMessageBox.warning(self, "Bazıları üretilemedi", message)
        else:
            QtWidgets.QMessageBox.information(self, "Toplu sertifika", message)
        self.state.status("%d sertifika üretildi, onay bekliyor." % len(made))

    def _compare(self):
        ids = self._selected_session_ids()
        if len(ids) < 2:
            QtWidgets.QMessageBox.information(
                self, "En az iki oturum",
                "Karşılaştırmak için Ctrl ya da Shift ile en az iki oturum "
                "seçin.")
            return
        from .compare_dialog import CompareDialog

        CompareDialog(ids, self).exec()

    # =====================================================================
    # Certificates
    # =====================================================================
    def _certificates_tab(self):
        w = QtWidgets.QWidget()
        lay = QtWidgets.QVBoxLayout(w)
        lay.setContentsMargins(0, 10, 0, 0)

        grid = QtWidgets.QGridLayout()
        self.cert_search = QtWidgets.QLineEdit()
        self.cert_search.setPlaceholderText(
            "Ara: sertifika no, seri no, model veya şirket")
        self.cert_search.textChanged.connect(self.reload_certificates)

        self.cert_dut_filter = QtWidgets.QComboBox()
        self.cert_dut_filter.currentIndexChanged.connect(self.reload_certificates)

        self.cert_state_filter = QtWidgets.QComboBox()
        cert_states = [("Tümü", None), ("Onay bekleyen", "pending"),
                       ("Onaylanmış", "approved")]
        if self.state.can(perms.CERT_VIEW_DELETED):
            cert_states.append(("Silinmiş", "deleted"))
        for label, value in cert_states:
            self.cert_state_filter.addItem(label, value)
        self.cert_state_filter.currentIndexChanged.connect(self.reload_certificates)

        self.cert_result_filter = QtWidgets.QComboBox()
        for label, value in (("Tüm sonuçlar", None), ("Uygun", "pass"),
                             ("Uygun değil", "fail"), ("Bilgilendirme", "info")):
            self.cert_result_filter.addItem(label, value)
        self.cert_result_filter.currentIndexChanged.connect(self.reload_certificates)

        self.cert_date_filter = DateRangeFilter()
        self.cert_date_filter.changed.connect(self.reload_certificates)

        grid.addWidget(self.cert_search, 0, 0, 1, 4)
        grid.addWidget(QtWidgets.QLabel("Cihaz"), 1, 0)
        grid.addWidget(self.cert_dut_filter, 1, 1)
        grid.addWidget(QtWidgets.QLabel("Durum"), 1, 2)
        grid.addWidget(self.cert_state_filter, 1, 3)
        grid.addWidget(QtWidgets.QLabel("Sonuç"), 1, 4)
        grid.addWidget(self.cert_result_filter, 1, 5)
        grid.addWidget(QtWidgets.QLabel("Tarih"), 2, 0)
        grid.addWidget(self.cert_date_filter, 2, 1, 1, 5)
        for c in (1, 3, 5):
            grid.setColumnStretch(c, 1)
        lay.addLayout(grid)

        self.cert_table = QtWidgets.QTableWidget(0, 8)
        self.cert_table.setHorizontalHeaderLabels(
            ["Sertifika no", "Tür", "Tarih", "Cihaz", "Şirket", "Sonuç",
             "Durum", "Silme gerekçesi"])
        self.cert_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.cert_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.cert_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.cert_table.setAlternatingRowColors(True)
        self.cert_table.verticalHeader().setVisible(False)
        self.cert_table.itemSelectionChanged.connect(self._on_cert_selected)
        self.cert_table.doubleClicked.connect(self._open_certificate)
        lay.addWidget(self.cert_table, 1)

        row = QtWidgets.QHBoxLayout()
        self.preview_btn = QtWidgets.QPushButton("Önizle")
        self.preview_btn.setToolTip(
            "Sertifikayı uygulama içinde gösterir — yanlış üretilmiş bir "
            "belgeyi fark etmek için\ndosyayı açıp kapatmak gerekmesin.")
        self.preview_btn.clicked.connect(self._preview_certificate)
        row.addWidget(self.preview_btn)
        self.open_cert_btn = QtWidgets.QPushButton("PDF'i aç")
        self.open_cert_btn.clicked.connect(self._open_certificate)
        self.zip_cert_btn = QtWidgets.QPushButton("Dosyaları zip indir")
        self.zip_cert_btn.setToolTip(
            "Seçili dalga serisindeki her şokun osiloskop ekran görüntüsünü "
            "(PNG) ve ham veri dosyasını (CSV) tek bir zip dosyasında "
            "indirir.")
        self.zip_cert_btn.clicked.connect(self._export_certificate_zip)
        self.approve_btn = QtWidgets.QPushButton("Onayla")
        self.approve_btn.setProperty("primary", True)
        self.approve_btn.clicked.connect(self._approve)
        self.delete_btn = QtWidgets.QPushButton("Kaydı sil")
        self.delete_btn.setProperty("danger", True)
        self.delete_btn.clicked.connect(self._delete_certificate)
        self.restore_btn = QtWidgets.QPushButton("Geri al")
        self.restore_btn.clicked.connect(self._restore_certificate)
        for b in (self.open_cert_btn, self.zip_cert_btn, self.approve_btn,
                  self.delete_btn, self.restore_btn):
            row.addWidget(b)
        # Hiding happens AFTER adding to the layout: a widget added to a
        # layout is shown again, so a setVisible(False) called beforehand
        # has no effect.
        self.approve_btn.setVisible(self.state.can(perms.CERT_APPROVE))
        self.delete_btn.setVisible(self.state.can(perms.CERT_DELETE))
        self.restore_btn.setVisible(self.state.can(perms.CERT_RESTORE))
        row.addStretch(1)
        lay.addLayout(row)

        self.cert_note = QtWidgets.QLabel()
        self.cert_note.setProperty("hint", True)
        self.cert_note.setWordWrap(True)
        lay.addWidget(self.cert_note)
        return w

    def _is_admin(self):
        return self.state.can(perms.CERT_VIEW_DELETED)

    def reload_certificates(self):
        previous = self.cert_dut_filter.currentData()
        self.cert_dut_filter.blockSignals(True)
        self.cert_dut_filter.clear()
        self.cert_dut_filter.addItem("Tüm cihazlar", None)
        for r in db.query(
            "SELECT d.id, d.manufacturer, d.model, d.serial_no FROM duts d"
            " WHERE EXISTS (SELECT 1 FROM certificates c"
            "               JOIN sessions s ON s.id = c.session_id"
            "               WHERE s.dut_id = d.id)"
            "    OR EXISTS (SELECT 1 FROM certificates c"
            "               JOIN waveform_captures w ON w.series_id = c.series_id"
            "               WHERE w.dut_id = d.id)"
            " ORDER BY d.manufacturer, d.model"
        ):
            self.cert_dut_filter.addItem(
                "%s %s — %s" % (r["manufacturer"], r["model"], r["serial_no"]),
                r["id"])
        idx = self.cert_dut_filter.findData(previous)
        self.cert_dut_filter.setCurrentIndex(idx if idx >= 0 else 0)
        self.cert_dut_filter.blockSignals(False)

        # The (certificate, device, kind) mapping lives in
        # `certificate.SOURCE_JOIN`: the approval queue uses the same
        # mapping, so the two screens don't end up with two different
        # answers to "which certificates exist".
        sql = (
            "SELECT c.*, k.dut_id, k.kind, d.manufacturer, d.model,"
            " d.serial_no, d.company, du.full_name AS deleted_by_name"
            " FROM certificates c" + certificate.SOURCE_JOIN +
            # See the note above certificate.pending(): for waveform series
            # without a dut_id, an INNER JOIN would drop the certificate
            # from the list.
            " LEFT JOIN duts d ON d.id = k.dut_id"
            " LEFT JOIN users du ON du.id = c.deleted_by"
        )
        where, params = [], []

        # Deleted records are visible only to admins
        if not self._is_admin():
            where.append("c.deleted_at IS NULL")

        term = self.cert_search.text().strip()
        if term:
            like = "%" + term + "%"
            where.append("(c.cert_no LIKE ? OR d.serial_no LIKE ? OR d.model LIKE ?"
                         " OR d.company LIKE ?)")
            params += [like, like, like, like]
        if self.cert_dut_filter.currentData() is not None:
            where.append("k.dut_id = ?")
            params.append(self.cert_dut_filter.currentData())
        state = self.cert_state_filter.currentData()
        if state == "pending":
            where.append("c.approved_at IS NULL AND c.deleted_at IS NULL")
        elif state == "approved":
            where.append("c.approved_at IS NOT NULL AND c.deleted_at IS NULL")
        elif state == "deleted":
            where.append("c.deleted_at IS NOT NULL")
        if self.cert_result_filter.currentData() is not None:
            where.append("c.result = ?")
            params.append(self.cert_result_filter.currentData())
        start, end = self.cert_date_filter.range()
        if start:
            where.append("c.issued_at >= ? AND c.issued_at < ?")
            params += [start, end]
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY c.id DESC LIMIT 500"

        c = theme.colors()
        rows = db.query(sql, tuple(params))
        self.cert_table.setRowCount(0)
        for r in rows:
            i = self.cert_table.rowCount()
            self.cert_table.insertRow(i)
            if r["deleted_at"]:
                state_text = "SİLİNDİ (%s)" % (r["deleted_by_name"] or "—")
            elif r["approved_at"]:
                state_text = "onaylandı"
            else:
                state_text = "onay bekliyor"
            cells = [
                r["cert_no"],
                certificate.KIND_TR.get(r["kind"], r["kind"]),
                (r["issued_at"] or "")[:10],
                certificate.device_label(r),
                r["company"] or "—", certificate.VERDICT_TR[r["result"]], state_text,
                r["delete_reason"] or "",
            ]
            for col, text in enumerate(cells):
                item = QtWidgets.QTableWidgetItem(str(text))
                if col == 0:
                    item.setData(Qt.UserRole, r["id"])
                if r["deleted_at"]:
                    item.setForeground(QtGui.QColor(c["text_muted"]))
                    font = item.font()
                    font.setStrikeOut(True)
                    item.setFont(font)
                self.cert_table.setItem(i, col, item)
        fit_table(self.cert_table, stretch_column=3)
        empty_state(self.cert_table, "Bu süzgece uyan sertifika yok.\nTarih aralığı: %s"
                    % self.cert_date_filter.describe())

        if self._is_admin():
            self.cert_note.setText(
                "Silinen sertifikalar listede üstü çizili görünür — kayıt "
                "veritabanından çıkmaz, numara serisinde boşluk oluşmaz.")
        else:
            self.cert_note.setText(
                "Silinen sertifikalar bu listede görünmez; yöneticiler görmeye "
                "devam eder.")
        self._on_cert_selected()

    def _selected_certificate(self):
        rows = self.cert_table.selectionModel().selectedRows()
        if not rows:
            return None
        cid = self.cert_table.item(rows[0].row(), 0).data(Qt.UserRole)
        return db.query_one("SELECT * FROM certificates WHERE id = ?", (cid,))

    def _on_cert_selected(self):
        cert = self._selected_certificate()
        deleted = bool(cert and cert["deleted_at"])
        can_manage = self.state.can(perms.CERT_APPROVE)

        self.open_cert_btn.setEnabled(bool(cert and cert["pdf_path"]))
        self.preview_btn.setEnabled(bool(cert and cert["pdf_path"]))
        self.zip_cert_btn.setEnabled(bool(cert and cert["session_id"] is None))
        self.approve_btn.setEnabled(
            bool(cert) and not deleted and not cert["approved_at"] and can_manage)
        self.delete_btn.setEnabled(bool(cert) and not deleted and can_manage)
        self.restore_btn.setEnabled(deleted and self._is_admin())

    def _open_certificate(self):
        cert = self._selected_certificate()
        if cert and cert["pdf_path"]:
            _open_file(cert["pdf_path"])

    def _preview_certificate(self):
        from . import pdf_preview

        cert = self._selected_certificate()
        if cert is None:
            return
        pdf_preview.show(cert["pdf_path"], self, title=cert["cert_no"])

    def _export_certificate_zip(self):
        """Downloads the screenshots and CSVs for the selected waveform series.

        Session certificates (multimeter readings) have no corresponding
        files on disk, so this is only enabled for waveform series
        certificates.
        """
        cert = self._selected_certificate()
        if cert is None or cert["session_id"] is not None:
            return
        rows = waveform.series_captures(cert["series_id"])
        files = []
        for r in rows:
            for key in ("file_path", "screenshot_path"):
                p = r[key]
                if p and os.path.isfile(p):
                    files.append(p)
        if not files:
            QtWidgets.QMessageBox.information(
                self, "Dosya yok",
                "Bu seriye ait diskte kayıtlı ekran görüntüsü ya da CSV "
                "bulunamadı.")
            return

        default_name = "%s-dosyalar.zip" % cert["cert_no"]
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "Zip olarak kaydet", default_name, "Zip (*.zip)")
        if not path:
            return

        try:
            with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
                seen = set()
                for p in files:
                    arcname = os.path.basename(p)
                    base, ext = os.path.splitext(arcname)
                    n = 1
                    while arcname in seen:
                        arcname = "%s_%d%s" % (base, n, ext)
                        n += 1
                    seen.add(arcname)
                    zf.write(p, arcname)
        except OSError as exc:
            QtWidgets.QMessageBox.warning(self, "Kaydedilemedi", str(exc))
            return

        audit.log("certificate.export_zip", user_id=self.state.user["id"],
                  entity="certificate", entity_id=cert["id"],
                  detail={"cert_no": cert["cert_no"], "path": path,
                          "count": len(files)})
        self.state.status(
            "%d dosya %s içine kaydedildi." % (len(files), os.path.basename(path)))
        QtWidgets.QMessageBox.information(
            self, "Tamamlandı",
            "%d dosya zip olarak kaydedildi:\n%s" % (len(files), path))

    def _approve(self):
        cert = self._selected_certificate()
        if cert is None:
            return
        ans = QtWidgets.QMessageBox.question(
            self, "Onay",
            "%s numaralı sertifika onaylansın mı?\n\n"
            "Onaydan sonra sertifika kilitlenir ve değiştirilemez."
            % cert["cert_no"],
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if ans != QtWidgets.QMessageBox.Yes:
            return
        try:
            certificate.approve(cert["id"], self.state.user["id"])
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Onaylanamadı", str(exc))
            return
        self.reload_certificates()
        self.state.status("Sertifika %s onaylandı." % cert["cert_no"])

    def _delete_certificate(self):
        cert = self._selected_certificate()
        if cert is None:
            return
        reason, ok = QtWidgets.QInputDialog.getText(
            self, "Kaydı sil",
            "%s silinecek.\n\nKayıt veritabanından çıkmaz, silinmiş olarak "
            "işaretlenir ve yöneticiler görmeye devam eder.\n\n"
            "Gerekçe (zorunlu):" % cert["cert_no"])
        if not ok:
            return
        try:
            certificate.soft_delete(cert["id"], self.state.user["id"], reason)
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Silinemedi", str(exc))
            return
        self.reload_certificates()
        self.state.status("%s silindi olarak işaretlendi." % cert["cert_no"])

    def _restore_certificate(self):
        cert = self._selected_certificate()
        if cert is None:
            return
        try:
            certificate.restore(cert["id"], self.state.user["id"])
        except (ValueError, PermissionError) as exc:
            QtWidgets.QMessageBox.warning(self, "Geri alınamadı", str(exc))
            return
        self.reload_certificates()
        self.state.status("%s geri alındı." % cert["cert_no"])


def _safe_filename(text, limit=90):
    """Strips characters that aren't valid in a filename.

    Session names are generated from the company name and can contain `/`
    or `:`; in a batch export that means the file fails to write at all.
    """
    cleaned = "".join(ch if ch not in '<>:"/\\|?*' else "-" for ch in text)
    cleaned = " ".join(cleaned.split()).strip(" .")
    return cleaned[:limit] or "oturum"


def _open_file(path):
    """Opens the file with the operating system's default application."""
    try:
        os.startfile(path)          # Windows
    except Exception:
        pass


def _write_excel(session_id, path):
    """Writes the session to Excel: summary + raw data sheets."""
    from openpyxl import Workbook

    d = certificate.collect(session_id)
    s, dut, inst = d["session"], d["dut"], d["instrument"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Özet"
    for row in (
        ("Oturum no", session_id),
        ("Ölçüm noktası sayısı", len(d["points"])),
        ("Oturum adı", (s["name"] or "").strip() or "—"),
        ("Tarih", s["started_at"]),
        ("Operatör", d["operator"]["full_name"]),
        ("Şirket", dut["company"]),
        ("Üretici", dut["manufacturer"]),
        ("Model", dut["model"]),
        ("Seri no", dut["serial_no"]),
        ("Referans cihaz", "%s %s (SN %s)" % (inst["brand"], inst["model"],
                                              inst["serial_no"])),
        ("Fonksiyon", s["function"]),
        ("Birim", s["unit"]),
        ("Okuma sayısı", d["n"]),
        ("Dışlanan", d["excluded"]),
        ("Nominal", d["nominal"]),
        ("Tolerans (±)", d["tolerance"]),
        ("Uygunluk kriteri", certificate.CRITERION_TR[d["mode"]]),
        ("Ortalama", d["mean"]),
        ("Std sapma", d["std"]),
        ("u (A tipi)", d["u_a"]),
        ("U (k=2)", d["U"]),
        ("En küçük", d["min"]),
        ("En büyük", d["max"]),
        ("Sapma", d["deviation"]),
        ("Sıcaklık (°C)", s["env_temp"]),
        ("Nem (%RH)", s["env_rh"]),
        ("Basınç (kPa)", s["env_pressure"]),
        ("Sonuç", certificate.VERDICT_TR[d["result"]]),
        ("Simülasyon", "EVET" if s["is_simulated"] else "hayır"),
    ):
        ws.append(list(row))

    # The points sheet is always written, even for a single-point session:
    # if whoever receives the file writes code that depends on the sheet's
    # presence, having it appear only sometimes would be the most annoying
    # surprise.
    ws_p = wb.create_sheet("Noktalar")
    ws_p.append(["#", "Fonksiyon", "Birim", "Nominal", "Tolerans (±)", "Kriter",
                 "n", "Dışlanan", "Ortalama", "Std sapma", "u (A tipi)",
                 "U (k=2)", "En küçük", "En büyük", "Sapma", "Sonuç"])
    for p in d["points"]:
        ws_p.append([p["seq"], p["function"], p["unit"], p["nominal"],
                     p["tolerance"], certificate.CRITERION_TR[p["mode"]],
                     p["n"], p["excluded"], p["mean"], p["std"], p["u_a"],
                     p["U"], p["min"], p["max"], p["deviation"],
                     certificate.VERDICT_TR[p["result"]]])

    ws2 = wb.create_sheet("Ham veri")
    ws2.append(["#", "Nokta", "Süre (s)", "Zaman (UTC)", "Değer", "Birim",
                "Ham yanıt", "Dışlandı", "Gerekçe"])
    seq_of = {p["point"]["id"]: p["seq"] for p in d["points"]}
    first_seq = d["points"][0]["seq"] if d["points"] else 1
    for r in db.query(
        "SELECT r.*, e.reason FROM readings r"
        " LEFT JOIN reading_exclusions e ON e.reading_id = r.id"
        " WHERE r.session_id = ? ORDER BY r.seq", (session_id,)
    ):
        # point_id NULL = first point (records predating the plan concept)
        point_seq = seq_of.get(r["point_id"], first_seq)
        ws2.append([r["seq"], point_seq, r["elapsed_s"], r["ts_utc"], r["value"],
                    r["unit"], r["raw"], "EVET" if r["reason"] else "",
                    r["reason"] or ""])

    wb.save(path)
